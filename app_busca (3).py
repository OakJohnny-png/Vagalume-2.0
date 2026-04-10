# =============================================================================
#  Stok Holmes — Sistema de Busca e Gestão de Devolução de Materiais
#  + Processador de Chamados de Iluminação Pública
#  + Solicitação de Materiais (Manual e via PDF)
#  Setor de Obras e Instalações Elétricas — NEMA
#  Desenvolvido com Python + Streamlit
# =============================================================================

import os
import re
import json
import glob
import io
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

try:
    import pdfplumber
    PDF_DISPONIVEL = True
except ImportError:
    PDF_DISPONIVEL = False

# ---------------------------------------------------------------------------
# Configuração global da página
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Stok Holmes",
    page_icon="🔍",
    layout="wide",
)

# ---------------------------------------------------------------------------
# Constantes e caminhos
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH        = os.path.join(BASE_DIR, "estoque.xlsx")
HISTORICO_DIR     = os.path.join(BASE_DIR, "historico_devolucoes")
REQUISICOES_DIR   = os.path.join(BASE_DIR, "historico_requisicoes")

# Arquivos de referência para selectboxes
REF_CLIENTES      = os.path.join(BASE_DIR, "clientes.xlsx")
REF_VENDEDORES    = os.path.join(BASE_DIR, "vendedores.xlsx")
REF_COMPRADORES   = os.path.join(BASE_DIR, "compradores.xlsx")
REF_TIPOS_VENDA   = os.path.join(BASE_DIR, "tipos_venda.xlsx")
REF_DEPARTAMENTOS = os.path.join(BASE_DIR, "departamentos.xlsx")

os.makedirs(HISTORICO_DIR, exist_ok=True)
os.makedirs(REQUISICOES_DIR, exist_ok=True)


# ---------------------------------------------------------------------------
# Funções auxiliares — STOK HOLMES
# ---------------------------------------------------------------------------

@st.cache_data
def carregar_estoque() -> pd.DataFrame:
    df = pd.read_excel(EXCEL_PATH, dtype=str)
    for col in ["Código", "Descrição", "Localização"]:
        if col not in df.columns:
            st.error(f"Coluna '{col}' não encontrada no arquivo estoque.xlsx.")
            st.stop()
    df["Código"] = df["Código"].str.strip()
    df["Descrição"] = df["Descrição"].str.strip()
    df["Localização"] = df["Localização"].str.strip()
    df.dropna(how="all", inplace=True)
    df.sort_values("Descrição", inplace=True, ignore_index=True)
    return df


@st.cache_data
def carregar_referencia(caminho: str, colunas: list) -> list:
    """Carrega um arquivo Excel de referência e retorna lista de strings formatadas."""
    if not os.path.exists(caminho):
        return []
    try:
        df = pd.read_excel(caminho, dtype=str).fillna("")
        resultado = []
        for _, row in df.iterrows():
            partes = [str(row[c]).strip() for c in colunas if c in df.columns and str(row[c]).strip()]
            if partes:
                resultado.append(" — ".join(partes) if len(partes) > 1 else partes[0])
        return resultado
    except Exception:
        return []


@st.cache_data
def carregar_clientes() -> list:
    """Retorna lista ['COD — NOME', ...] do arquivo clientes.xlsx."""
    return carregar_referencia(REF_CLIENTES, ["Código", "Nome"])


@st.cache_data
def carregar_vendedores() -> list:
    return carregar_referencia(REF_VENDEDORES, ["Código", "Nome"])


@st.cache_data
def carregar_compradores() -> list:
    return carregar_referencia(REF_COMPRADORES, ["Nome"])


@st.cache_data
def carregar_tipos_venda() -> list:
    return carregar_referencia(REF_TIPOS_VENDA, ["Tipo"])


@st.cache_data
def carregar_departamentos() -> list:
    return carregar_referencia(REF_DEPARTAMENTOS, ["Código", "Nome"])


def gerar_protocolo() -> str:
    hoje = datetime.now().strftime("%d%m%Y")
    padrao = os.path.join(HISTORICO_DIR, f"DEV_{hoje}*.json")
    arquivos_hoje = glob.glob(padrao)
    sequencial = len(arquivos_hoje) + 1
    return f"{hoje}.{sequencial:03d}"


def gerar_numero_solicitacao(cliente_cod: str) -> str:
    """
    Gera número de solicitação no formato: CODddmmAAAASS
    onde SS é sequencial único por cliente no dia.
    Exemplo: cliente 1090, dia 09/04/2026, 1ª solicitação → 10900904202601
    """
    hoje = datetime.now().strftime("%d%m%Y")
    cod = re.sub(r'\D', '', str(cliente_cod))  # apenas dígitos do código
    prefixo = f"{cod}{hoje}"
    # Conta quantos arquivos já existem para esse prefixo hoje
    padrao = os.path.join(REQUISICOES_DIR, f"SOL_{prefixo}*.json")
    existentes = glob.glob(padrao)
    sequencial = len(existentes) + 1
    return f"{prefixo}{sequencial:02d}"


def df_lista_vazio() -> pd.DataFrame:
    return pd.DataFrame(columns=["Código", "Descrição", "Localização", "Quantidade"])


def inicializar_session_state():
    if "lista_devolucao" not in st.session_state:
        st.session_state.lista_devolucao = df_lista_vazio()
    if "empresa" not in st.session_state:
        st.session_state.empresa = ""
    if "protocolo" not in st.session_state:
        st.session_state.protocolo = gerar_protocolo()
    if "req_cabecalho" not in st.session_state:
        st.session_state.req_cabecalho = {}
    if "req_itens" not in st.session_state:
        st.session_state.req_itens = []
    if "req_manual_itens" not in st.session_state:
        st.session_state.req_manual_itens = []
    if "req_manual_cab" not in st.session_state:
        st.session_state.req_manual_cab = {}


def salvar_devolucao():
    protocolo = st.session_state.protocolo
    dados = {
        "protocolo": protocolo,
        "empresa": st.session_state.empresa,
        "data_hora": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "itens": st.session_state.lista_devolucao.to_dict(orient="records"),
    }
    caminho = os.path.join(HISTORICO_DIR, f"DEV_{protocolo}.json")
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=4)
    st.success(f"Devolução salva com sucesso! Protocolo: **{protocolo}**")


def nova_devolucao():
    st.session_state.lista_devolucao = df_lista_vazio()
    st.session_state.empresa = ""
    st.session_state.protocolo = gerar_protocolo()


def exportar_excel() -> bytes:
    df_export = st.session_state.lista_devolucao.copy()
    df_export.insert(0, "Protocolo", st.session_state.protocolo)
    df_export.insert(1, "Empresa", st.session_state.empresa)
    df_export["Data/Hora"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Devolução")
    return buffer.getvalue()


def listar_historico() -> list:
    arquivos = sorted(glob.glob(os.path.join(HISTORICO_DIR, "DEV_*.json")), reverse=True)
    registros = []
    for arq in arquivos:
        with open(arq, "r", encoding="utf-8") as f:
            dados = json.load(f)
        registros.append({
            "arquivo": arq,
            "protocolo": dados.get("protocolo", "—"),
            "empresa": dados.get("empresa", "—"),
            "data_hora": dados.get("data_hora", "—"),
            "qtd_itens": len(dados.get("itens", [])),
            "itens": dados.get("itens", []),
        })
    return registros


def carregar_historico_na_sessao(registro: dict):
    st.session_state.protocolo = registro["protocolo"]
    st.session_state.empresa = registro["empresa"]
    itens = registro["itens"]
    if itens:
        df = pd.DataFrame(itens)
        df["Quantidade"] = pd.to_numeric(df["Quantidade"], errors="coerce").fillna(0).astype(int)
        st.session_state.lista_devolucao = df
    else:
        st.session_state.lista_devolucao = df_lista_vazio()


# ---------------------------------------------------------------------------
# Funções auxiliares — PROCESSADOR DE CHAMADOS IP
# ---------------------------------------------------------------------------

ROUTES = {
    "ROTA 1": ["CENTRO", "JARDIM AMERICA"],
    "ROTA 2": ["ALBERTINA", "LARANJEIRAS", "BOA VISTA", "EUGENIO SCHNEIDER"],
    "ROTA 3": ["FUNDO CANOAS", "CANOAS", "PROGRESSO", "PAMPLONA", "CANTA GALO"],
    "ROTA 4": ["BARRA DO TROMBUDO", "BARRAGEM", "BUDAG", "SUMARE"],
    "ROTA 5": ["SANTANA", "TABOAO", "BREMER", "BELA ALIANÇA"],
    "ROTA 6": ["BARRA DA ITOUPAVA", "NAVEGANTES", "SANTA RITA", "VALADA ITOUPAVA", "VALADA SÃO PAULO", "RAINHA"]
}


def processar_chamados(
    uploaded_file, fonte_escolhida,
    cor_fundo_rota, cor_fonte_rota, tamanho_rota,
    cor_fundo_bairro, cor_fonte_bairro, tamanho_bairro,
    cor_fundo_prob, cor_fonte_prob, tamanho_prob,
) -> bytes:
    xls = pd.read_excel(uploaded_file, sheet_name=None, header=None, dtype=str)
    abas_disponiveis = {nome.strip().upper(): nome for nome in xls.keys()}
    data_by_route = {r: {} for r in ROUTES}

    for route, neighborhoods in ROUTES.items():
        for neighborhood in neighborhoods:
            nome_aba_upper = neighborhood.upper()
            if nome_aba_upper in abas_disponiveis:
                nome_real_aba = abas_disponiveis[nome_aba_upper]
                df = xls[nome_real_aba]
                if len(df.columns) >= 4:
                    df_filtered = df[
                        (df[3].isin(['NÃO REALIZADO', 'NÃO EXECUTADO'])) &
                        (df[1].notna()) & (df[1].str.strip() != "")
                    ]
                    problems = df_filtered[1].tolist()
                    if problems:
                        data_by_route[route][neighborhood] = problems

    wb = Workbook()
    ws = wb.active
    ws.title = "Chamados Pendentes"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    hex_bg_rota   = cor_fundo_rota.replace('#', '')
    hex_fg_rota   = cor_fonte_rota.replace('#', '')
    hex_bg_bairro = cor_fundo_bairro.replace('#', '')
    hex_fg_bairro = cor_fonte_bairro.replace('#', '')
    hex_bg_prob   = cor_fundo_prob.replace('#', '')
    hex_fg_prob   = cor_fonte_prob.replace('#', '')

    font_rota   = Font(name=fonte_escolhida, size=tamanho_rota,   color=hex_fg_rota,   bold=True)
    fill_rota   = PatternFill(start_color=hex_bg_rota,   end_color=hex_bg_rota,   fill_type="solid")
    font_bairro = Font(name=fonte_escolhida, size=tamanho_bairro, color=hex_fg_bairro, bold=True)
    fill_bairro = PatternFill(start_color=hex_bg_bairro, end_color=hex_bg_bairro, fill_type="solid")
    font_prob   = Font(name=fonte_escolhida, size=tamanho_prob,   color=hex_fg_prob)
    fill_prob   = PatternFill(start_color=hex_bg_prob,   end_color=hex_bg_prob,   fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    current_row = 1
    for route, neighborhoods in data_by_route.items():
        if not neighborhoods:
            continue
        cell = ws.cell(row=current_row, column=1, value=route)
        cell.font = font_rota; cell.fill = fill_rota
        cell.border = thin_border; cell.alignment = Alignment(wrap_text=True)
        current_row += 1
        first_bairro = True
        for bairro, problems in neighborhoods.items():
            if not first_bairro:
                current_row += 1
            first_bairro = False
            cell = ws.cell(row=current_row, column=1, value=bairro)
            cell.font = font_bairro; cell.fill = fill_bairro
            cell.border = thin_border; cell.alignment = Alignment(wrap_text=True)
            current_row += 1
            for problem in problems:
                cell = ws.cell(row=current_row, column=1, value=str(problem).strip())
                cell.font = font_prob
                if hex_bg_prob != "FFFFFF":
                    cell.fill = fill_prob
                cell.border = thin_border; cell.alignment = Alignment(wrap_text=True)
                current_row += 1

    ws.column_dimensions['A'].width = 150
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ---------------------------------------------------------------------------
# Funções auxiliares — SOLICITAÇÃO DE MATERIAIS
# ---------------------------------------------------------------------------

def extrair_dados_requisicao(pdf_bytes) -> dict:
    """
    Extrai dados do cabeçalho e itens de uma requisição de materiais em PDF.
    Retorna um dicionário com 'cabecalho' e 'itens'.
    Suporta tanto unidades separadas por espaço quanto coladas à descrição.
    Pedido Original e Revisão são ignorados (não extraídos).
    """
    cabecalho = {}
    itens = []

    with pdfplumber.open(pdf_bytes) as pdf:
        texto_completo = ""
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                texto_completo += t + "\n"

    linhas = texto_completo.split("\n")

    for linha in linhas:
        # Número do Orçamento (referência do PDF)
        m = re.search(r'Orçamento[:\s]+(\d+)', linha)
        if m and 'orcamento_pdf' not in cabecalho:
            cabecalho['orcamento_pdf'] = m.group(1).strip()

        # Cliente (código + nome)
        m = re.search(r'Cliente[:\s]+(\d+)\s+(.+?)(?=\s{2,}|Comprador|$)', linha)
        if m and 'cliente_nome' not in cabecalho:
            cabecalho['cliente_cod'] = m.group(1).strip()
            cabecalho['cliente_nome'] = m.group(2).strip()

        # Vendedor
        m = re.search(r'Vendedor[.\s:]+(.+?)(?=\s{2,}|Prazo|$)', linha)
        if m and 'vendedor' not in cabecalho:
            cabecalho['vendedor'] = m.group(1).strip()

        # Comprador
        m = re.search(r'Comprador[.\s:]+(.+?)(?=\s{2,}|Em\.\.|$)', linha)
        if m and 'comprador' not in cabecalho:
            val = re.sub(r'Em\.+:\s*\S+', '', m.group(1)).strip()
            if val:
                cabecalho['comprador'] = val

        # Tipo de Venda
        m = re.search(r'Tipo de Venda[.\s:]+(.+)', linha)
        if m and 'tipo_venda' not in cabecalho:
            cabecalho['tipo_venda'] = m.group(1).strip()

        # Departamento
        m = re.search(r'Departamento[.\s:]+(.+)', linha)
        if m and 'departamento' not in cabecalho:
            cabecalho['departamento'] = m.group(1).strip()

        # Marcações
        m = re.search(r'Marcações[.\s:]+(.+)', linha)
        if m and 'marcacoes' not in cabecalho:
            val = m.group(1).strip().lstrip(':').strip()
            cabecalho['marcacoes'] = val if val else "—"

        # Observação
        m = re.search(r'Observação\s*:\s*(.+)', linha)
        if m and 'observacao' not in cabecalho:
            cabecalho['observacao'] = m.group(1).strip()

    # --- Extrair itens da tabela ---
    _UN = r'PC|KG|MT|CX|PAR|VB|SC|BD|GL|TB|KIT|FD|CJ|JG|RL|LT|UN|M'

    padrao_normal = re.compile(
        r'^(\d{2})\s+(\d{4,6})\s+(.+?)\s+(' + _UN + r')\s+([\d.,]+)',
        re.IGNORECASE
    )
    padrao_colado = re.compile(
        r'^(\d{2})\s+(\d{4,6})\s+(.+?)(' + _UN + r')\s+([\d.,]+)$',
        re.IGNORECASE
    )

    for linha in linhas:
        linha_strip = linha.strip()
        m = padrao_normal.match(linha_strip)
        if m:
            seq, codigo = m.group(1), m.group(2)
            descricao = m.group(3).strip()
            unidade = m.group(4).upper()
            try:
                quantidade = float(m.group(5).replace(',', '.'))
            except ValueError:
                quantidade = 0.0
            itens.append({
                "Seq": seq, "Localização": "", "Código": codigo,
                "Descrição": descricao, "UN": unidade, "Quantidade": quantidade,
            })
            continue

        m = padrao_colado.match(linha_strip)
        if m:
            seq, codigo = m.group(1), m.group(2)
            descricao = m.group(3).strip()
            unidade = m.group(4).upper()
            try:
                quantidade = float(m.group(5).replace(',', '.'))
            except ValueError:
                quantidade = 0.0
            itens.append({
                "Seq": seq, "Localização": "", "Código": codigo,
                "Descrição": descricao, "UN": unidade, "Quantidade": quantidade,
            })

    return {"cabecalho": cabecalho, "itens": itens}


def salvar_requisicao(cabecalho: dict, itens: list) -> str:
    """
    Salva a solicitação em JSON. O número de solicitação é gerado automaticamente
    no formato CODddmmAAAASS e armazenado em cabecalho['num_solicitacao'].
    Retorna o nome do arquivo salvo.
    """
    cliente_cod = cabecalho.get('cliente_cod', '0')
    num_sol = gerar_numero_solicitacao(cliente_cod)
    cabecalho['num_solicitacao'] = num_sol

    nome_arquivo = f"SOL_{num_sol}.json"
    caminho = os.path.join(REQUISICOES_DIR, nome_arquivo)
    dados = {
        "cabecalho": cabecalho,
        "itens": itens,
        "data_importacao": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    }
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=4)
    return nome_arquivo


def listar_requisicoes() -> list:
    """
    Lê todos os arquivos JSON de solicitações e retorna lista de dicionários,
    ordenada por data (mais recente primeiro).
    """
    arquivos = sorted(
        glob.glob(os.path.join(REQUISICOES_DIR, "SOL_*.json")),
        reverse=True,
    )
    registros = []
    for arq in arquivos:
        with open(arq, "r", encoding="utf-8") as f:
            dados = json.load(f)
        cab = dados.get("cabecalho", {})
        registros.append({
            "arquivo": arq,
            "num_solicitacao": cab.get("num_solicitacao", "—"),
            "orcamento_pdf": cab.get("orcamento_pdf", "—"),
            "cliente_cod": cab.get("cliente_cod", "—"),
            "cliente_nome": cab.get("cliente_nome", "—"),
            "vendedor": cab.get("vendedor", "—"),
            "departamento": cab.get("departamento", "—"),
            "observacao": cab.get("observacao", "—"),
            "data_importacao": dados.get("data_importacao", "—"),
            "qtd_itens": len(dados.get("itens", [])),
            "cabecalho": cab,
            "itens": dados.get("itens", []),
        })
    return registros


def exportar_requisicao_excel(cabecalho: dict, itens: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitação"

    fonte_titulo    = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    fill_titulo     = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fonte_cab_label = Font(name="Calibri", size=10, bold=True)
    fonte_cab_valor = Font(name="Calibri", size=10)
    fill_cab        = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fonte_header    = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    fill_header     = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    fonte_item      = Font(name="Calibri", size=10)
    fill_item_alt   = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    borda = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    alinhamento_centro = Alignment(horizontal="center", vertical="center")
    alinhamento_esq    = Alignment(horizontal="left", vertical="center", wrap_text=True)

    linha = 1

    # Título
    ws.merge_cells(f"A{linha}:G{linha}")
    cell = ws.cell(row=linha, column=1, value="SOLICITAÇÃO DE MATERIAIS — NEMA")
    cell.font = fonte_titulo; cell.fill = fill_titulo
    cell.alignment = alinhamento_centro
    ws.row_dimensions[linha].height = 22
    linha += 1

    # Cabeçalho
    campos_cab = [
        ("Nº Solicitação", cabecalho.get("num_solicitacao", "—")),
        ("Nº Orçamento PDF", cabecalho.get("orcamento_pdf", "—")),
        ("Tipo de Pedido", cabecalho.get("tipo_pedido", "—")),
        ("Cliente", f"{cabecalho.get('cliente_cod', '')} — {cabecalho.get('cliente_nome', '')}"),
        ("Vendedor", cabecalho.get("vendedor", "—")),
        ("Comprador", cabecalho.get("comprador", "—")),
        ("Tipo de Venda", cabecalho.get("tipo_venda", "—")),
        ("Departamento", cabecalho.get("departamento", "—")),
        ("Marcações", cabecalho.get("marcacoes", "—")),
        ("Observação", cabecalho.get("observacao", "—")),
    ]

    for label, valor in campos_cab:
        ws.merge_cells(f"A{linha}:B{linha}")
        cell_label = ws.cell(row=linha, column=1, value=label)
        cell_label.font = fonte_cab_label; cell_label.fill = fill_cab
        cell_label.border = borda; cell_label.alignment = alinhamento_esq

        ws.merge_cells(f"C{linha}:G{linha}")
        cell_valor = ws.cell(row=linha, column=3, value=valor)
        cell_valor.font = fonte_cab_valor
        cell_valor.border = borda; cell_valor.alignment = alinhamento_esq
        linha += 1

    linha += 1  # Espaço

    # Cabeçalho da tabela de itens
    headers     = ["Seq", "Local.", "Código", "Descrição", "UN", "Quantidade"]
    col_widths  = [6, 14, 12, 52, 8, 12]
    col_letters = ["A", "B", "C", "D", "E", "F"]

    for i, h in enumerate(headers):
        cell = ws.cell(row=linha, column=i + 1, value=h)
        cell.font = fonte_header; cell.fill = fill_header
        cell.border = borda; cell.alignment = alinhamento_centro
    ws.row_dimensions[linha].height = 18
    linha += 1

    # Itens
    for idx, item in enumerate(itens):
        fill_atual = fill_item_alt if idx % 2 == 0 else None
        valores = [
            item.get("Seq", ""),
            item.get("Localização", ""),
            item.get("Código", ""),
            item.get("Descrição", ""),
            item.get("UN", ""),
            item.get("Quantidade", 0),
        ]
        for i, val in enumerate(valores):
            cell = ws.cell(row=linha, column=i + 1, value=val)
            cell.font = fonte_item; cell.border = borda
            cell.alignment = alinhamento_esq if i == 3 else alinhamento_centro
            if fill_atual:
                cell.fill = fill_atual
        ws.row_dimensions[linha].height = 15
        linha += 1

    for col_letter, width in zip(col_letters, col_widths):
        ws.column_dimensions[col_letter].width = width
    ws.column_dimensions["G"].width = 5

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ---------------------------------------------------------------------------
# Inicialização
# ---------------------------------------------------------------------------
inicializar_session_state()
df_estoque = carregar_estoque()

# Carrega listas de referência
lista_clientes     = carregar_clientes()
lista_vendedores   = carregar_vendedores()
lista_compradores  = carregar_compradores()
lista_tipos_venda  = carregar_tipos_venda()
lista_departamentos = carregar_departamentos()

# ---------------------------------------------------------------------------
# NAVEGAÇÃO SUPERIOR — Menu em Abas
# ---------------------------------------------------------------------------
st.markdown("""
    <style>
        .stTabs [data-baseweb="tab-list"] button [data-testid="stMarkdownContainer"] p {
            font-size: 1.1rem;
            font-weight: 600;
        }
    </style>
""", unsafe_allow_html=True)

tab_inicio, tab_finder, tab_devolucao, tab_historico, tab_chamados, tab_requisicoes = st.tabs(
    ["🏠 Início", "🔍 Finder", "📦 Devolução", "📂 Histórico", "💡 Chamados IP", "📋 Solicitação de Materiais"]
)

# ===========================================================================
# ABA 1: INÍCIO
# ===========================================================================
with tab_inicio:
    st.markdown(
        "<h3 style='text-align:center; color:gray;'>Bem vindo ao</h3>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<h1 style='text-align:center; font-size:85px; color:#4CAF50;'>Stok Holmes</h1>",
        unsafe_allow_html=True,
    )
    st.markdown("---")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total de itens no estoque", len(df_estoque))
    with col2:
        st.metric("Protocolo atual", st.session_state.protocolo)
    with col3:
        st.metric("Itens na devolução atual", len(st.session_state.lista_devolucao))

    st.markdown("---")
    st.info(
        "Use as abas acima para navegar entre as funcionalidades do sistema.\n\n"
        "- **Finder**: pesquise itens do estoque por código ou descrição.\n"
        "- **Devolução**: monte e exporte listas de devolução de materiais.\n"
        "- **Histórico**: consulte e reabra devoluções anteriores.\n"
        "- **Chamados IP**: processe planilhas de chamados de iluminação pública.\n"
        "- **Solicitação de Materiais**: crie solicitações manualmente ou importe a partir de arquivos PDF."
    )


# ===========================================================================
# ABA 2: FINDER
# ===========================================================================
with tab_finder:
    st.title("🔍 Finder — Busca de Materiais")
    st.markdown("Pesquise itens pelo **Código** ou pela **Descrição**.")

    termo = st.text_input("Digite o código ou a descrição do item:", placeholder="Ex: 1001 ou CABO")

    if termo.strip():
        mask = (
            df_estoque["Código"].str.contains(termo.strip(), case=False, na=False)
            | df_estoque["Descrição"].str.contains(termo.strip(), case=False, na=False)
        )
        resultado = df_estoque[mask].sort_values("Descrição").reset_index(drop=True)

        if resultado.empty:
            st.warning("Nenhum item encontrado para o termo pesquisado.")
        else:
            st.success(f"{len(resultado)} item(ns) encontrado(s).")
            st.dataframe(resultado, use_container_width=True, hide_index=True)
    else:
        st.dataframe(df_estoque, use_container_width=True, hide_index=True)


# ===========================================================================
# ABA 3: DEVOLUÇÃO
# ===========================================================================
with tab_devolucao:
    st.title("📦 Devolução de Materiais")
    st.markdown(f"**Protocolo:** `{st.session_state.protocolo}`")

    empresa_input = st.text_input(
        "Nome da Empresa que está devolvendo:",
        value=st.session_state.empresa,
        key="empresa_input",
    )
    st.session_state.empresa = empresa_input

    st.markdown("---")

    col_esq, col_dir = st.columns([1.2, 1])

    with col_esq:
        st.subheader("Adicionar Item")

        busca_item = st.text_input(
            "Buscar item (código ou descrição):",
            key="busca_item",
            placeholder="Ex: DISJUNTOR",
        )

        opcoes_encontradas = pd.DataFrame()
        if busca_item.strip():
            mask = (
                df_estoque["Código"].str.contains(busca_item.strip(), case=False, na=False)
                | df_estoque["Descrição"].str.contains(busca_item.strip(), case=False, na=False)
            )
            opcoes_encontradas = df_estoque[mask].sort_values("Descrição").reset_index(drop=True)

        item_selecionado = None
        if not opcoes_encontradas.empty:
            rotulos = [
                f"{row['Código']} — {row['Descrição']}"
                for _, row in opcoes_encontradas.iterrows()
            ]
            escolha = st.selectbox("Selecione o item:", rotulos, key="selectbox_item")
            idx = rotulos.index(escolha)
            item_selecionado = opcoes_encontradas.iloc[idx]
        elif busca_item.strip():
            st.warning("Nenhum item encontrado.")

        quantidade = st.number_input(
            "Quantidade:", min_value=1, value=1, step=1, key="quantidade_input",
        )

        if st.button("➕ Adicionar à lista", use_container_width=True):
            if item_selecionado is None:
                st.error("Selecione um item antes de adicionar.")
            elif not st.session_state.empresa.strip():
                st.error("Informe o nome da empresa antes de adicionar itens.")
            else:
                codigo = item_selecionado["Código"]
                lista = st.session_state.lista_devolucao

                if codigo in lista["Código"].values:
                    idx_existente = lista.index[lista["Código"] == codigo][0]
                    lista.at[idx_existente, "Quantidade"] = (
                        int(lista.at[idx_existente, "Quantidade"]) + int(quantidade)
                    )
                    st.success(f"Quantidade atualizada para o item **{item_selecionado['Descrição']}**.")
                else:
                    nova_linha = pd.DataFrame([{
                        "Código": codigo,
                        "Descrição": item_selecionado["Descrição"],
                        "Localização": item_selecionado["Localização"],
                        "Quantidade": int(quantidade),
                    }])
                    lista = pd.concat([lista, nova_linha], ignore_index=True)
                    st.success(f"Item **{item_selecionado['Descrição']}** adicionado.")

                lista.sort_values("Descrição", inplace=True, ignore_index=True)
                st.session_state.lista_devolucao = lista

        st.markdown("---")

        col_salvar, col_nova = st.columns(2)
        with col_salvar:
            if st.button("💾 Salvar esta Devolução", use_container_width=True):
                if st.session_state.lista_devolucao.empty:
                    st.error("A lista de devolução está vazia.")
                elif not st.session_state.empresa.strip():
                    st.error("Informe o nome da empresa antes de salvar.")
                else:
                    salvar_devolucao()

        with col_nova:
            if st.button("🆕 Nova Devolução", use_container_width=True):
                nova_devolucao()
                st.rerun()

    with col_dir:
        st.subheader("Lista de Devolução Atual")

        if st.session_state.lista_devolucao.empty:
            st.info("Nenhum item adicionado ainda.")
        else:
            st.dataframe(
                st.session_state.lista_devolucao,
                use_container_width=True,
                hide_index=True,
            )

            nome_arquivo = f"DEV_{st.session_state.protocolo}.xlsx"
            dados_excel = exportar_excel()
            st.download_button(
                label="📥 Exportar Excel",
                data=dados_excel,
                file_name=nome_arquivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )


# ===========================================================================
# ABA 4: HISTÓRICO / ABRIR
# ===========================================================================
with tab_historico:
    st.title("📂 Histórico de Devoluções")

    registros = listar_historico()

    if not registros:
        st.info("Nenhuma devolução salva ainda. Salve uma devolução na aba **Devolução**.")
    else:
        st.markdown(f"**{len(registros)}** registro(s) encontrado(s).")
        st.markdown("---")

        for reg in registros:
            with st.expander(
                f"📄 Protocolo: {reg['protocolo']}  |  Empresa: {reg['empresa']}  |  "
                f"Itens: {reg['qtd_itens']}  |  Data: {reg['data_hora']}"
            ):
                if reg["itens"]:
                    df_hist = pd.DataFrame(reg["itens"])
                    st.dataframe(df_hist, use_container_width=True, hide_index=True)
                else:
                    st.write("Lista de itens vazia.")

                if st.button(f"📂 Abrir protocolo {reg['protocolo']}", key=reg["protocolo"]):
                    carregar_historico_na_sessao(reg)
                    st.success(
                        f"Devolução **{reg['protocolo']}** carregada na sessão. "
                        "Acesse a aba **Devolução** para continuar o trabalho."
                    )


# ===========================================================================
# ABA 5: PROCESSADOR DE CHAMADOS IP
# ===========================================================================
with tab_chamados:
    st.title("💡 Processador de Chamados de Iluminação Pública")
    st.write("Faça o upload da planilha Excel contendo as abas dos bairros e personalize a formatação do arquivo final.")

    with st.expander("🎨 Personalização de Formatação", expanded=True):
        st.subheader("Configurações de Estilo")

        fonte_escolhida = st.selectbox(
            "Estilo da Fonte",
            ["Helvetica", "Arial", "Calibri", "Times New Roman", "Tahoma"],
            key="fonte_chamados"
        )

        st.subheader("1. Formatação das Rotas")
        col_rota1, col_rota2 = st.columns(2)
        with col_rota1:
            cor_fundo_rota = st.color_picker("Cor de Fundo (Rotas)", "#FF0000", key="bg_rota")
        with col_rota2:
            cor_fonte_rota = st.color_picker("Cor da Fonte (Rotas)", "#000000", key="fg_rota")
        tamanho_rota = st.number_input("Tamanho da Fonte (Rotas)", min_value=8, max_value=36, value=16, key="size_rota")

        st.subheader("2. Formatação dos Bairros")
        col_bairro1, col_bairro2 = st.columns(2)
        with col_bairro1:
            cor_fundo_bairro = st.color_picker("Cor de Fundo (Bairros)", "#D3D3D3", key="bg_bairro")
        with col_bairro2:
            cor_fonte_bairro = st.color_picker("Cor da Fonte (Bairros)", "#000000", key="fg_bairro")
        tamanho_bairro = st.number_input("Tamanho da Fonte (Bairros)", min_value=8, max_value=36, value=16, key="size_bairro")

        st.subheader("3. Formatação dos Problemas")
        col_prob1, col_prob2 = st.columns(2)
        with col_prob1:
            cor_fundo_prob = st.color_picker("Cor de Fundo (Problemas)", "#FFFFFF", key="bg_prob")
        with col_prob2:
            cor_fonte_prob = st.color_picker("Cor da Fonte (Problemas)", "#000000", key="fg_prob")
        tamanho_prob = st.number_input("Tamanho da Fonte (Problemas)", min_value=8, max_value=36, value=14, key="size_prob")

    st.markdown("---")
    uploaded_file = st.file_uploader("Arraste e solte sua planilha Excel (.xlsx) aqui", type=["xlsx"], key="upload_chamados")

    if uploaded_file is not None:
        if st.button("Processar Planilha", key="btn_processar"):
            with st.spinner('Lendo e processando os dados... Isso pode levar alguns segundos.'):
                try:
                    dados_excel = processar_chamados(
                        uploaded_file, fonte_escolhida,
                        cor_fundo_rota, cor_fonte_rota, tamanho_rota,
                        cor_fundo_bairro, cor_fonte_bairro, tamanho_bairro,
                        cor_fundo_prob, cor_fonte_prob, tamanho_prob,
                    )
                    st.success("✅ Processamento concluído com sucesso!")
                    st.download_button(
                        label="📥 Baixar Planilha Pronta",
                        data=dados_excel,
                        file_name="Chamados_Pendentes_Rotas.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_chamados"
                    )
                except Exception as e:
                    st.error(f"❌ Ocorreu um erro ao processar o arquivo: {e}")


# ===========================================================================
# ABA 6: SOLICITAÇÃO DE MATERIAIS
# ===========================================================================
with tab_requisicoes:
    st.title("📋 Solicitação de Materiais")

    # Três sub-abas
    sub_manual, sub_pdf, sub_historico = st.tabs([
        "✏️ Criar Manualmente",
        "📄 Importar via PDF",
        "📂 Solicitações",
    ])

    # =======================================================================
    # SUB-ABA: CRIAR MANUALMENTE
    # =======================================================================
    with sub_manual:
        st.subheader("✏️ Nova Solicitação Manual")
        st.markdown("Preencha os dados abaixo. Código, Descrição e Localização são buscados automaticamente do estoque.")
        st.markdown("---")

        # --- CABEÇALHO MANUAL ---
        st.subheader("📌 Dados do Cabeçalho")
        col_m1, col_m2 = st.columns(2)

        with col_m1:
            # Tipo de Pedido (substitui Pedido Original)
            man_tipo_pedido = st.selectbox(
                "Tipo de Pedido",
                ["", "Orçamento", "Pedido", "Devolução"],
                key="man_tipo_pedido"
            )

            # Cliente via selectbox (arquivo clientes.xlsx)
            opcoes_cliente = [""] + lista_clientes
            man_cliente_sel = st.selectbox("Cliente", opcoes_cliente, key="man_cliente_sel")
            # Extrai código e nome da seleção "COD — NOME"
            if man_cliente_sel and " — " in man_cliente_sel:
                man_cliente_cod = man_cliente_sel.split(" — ")[0].strip()
                man_cliente_nome = man_cliente_sel.split(" — ", 1)[1].strip()
            else:
                man_cliente_cod = ""
                man_cliente_nome = man_cliente_sel

            # Comprador via selectbox (arquivo compradores.xlsx)
            opcoes_comprador = [""] + lista_compradores
            man_comprador = st.selectbox("Comprador", opcoes_comprador, key="man_comprador")

        with col_m2:
            # Vendedor via selectbox (arquivo vendedores.xlsx)
            opcoes_vendedor = [""] + lista_vendedores
            man_vendedor_sel = st.selectbox("Vendedor", opcoes_vendedor, key="man_vendedor_sel")

            # Tipo de Venda via selectbox (arquivo tipos_venda.xlsx)
            opcoes_tipo_venda = [""] + lista_tipos_venda
            man_tipo_venda = st.selectbox("Tipo de Venda", opcoes_tipo_venda, key="man_tipo_venda")

            # Departamento via selectbox (arquivo departamentos.xlsx)
            opcoes_depto = [""] + lista_departamentos
            man_departamento = st.selectbox("Departamento", opcoes_depto, key="man_departamento")

            man_marcacoes = st.text_input("Marcações", key="man_marcacoes")
            man_observacao = st.text_input("Observação", key="man_observacao")

        st.markdown("---")

        # --- ADICIONAR ITENS ---
        st.subheader("📦 Adicionar Itens")
        col_busca, col_qtd, col_un = st.columns([3, 1, 1])

        with col_busca:
            man_busca = st.text_input(
                "Buscar item no estoque (código ou descrição):",
                key="man_busca_item",
                placeholder="Ex: CABO ou 160001"
            )

        man_item_sel = None
        if man_busca.strip():
            mask_man = (
                df_estoque["Código"].str.contains(man_busca.strip(), case=False, na=False)
                | df_estoque["Descrição"].str.contains(man_busca.strip(), case=False, na=False)
            )
            man_opcoes = df_estoque[mask_man].sort_values("Descrição").reset_index(drop=True)

            if not man_opcoes.empty:
                rotulos_man = [
                    f"{row['Código']} — {row['Descrição']} [{row['Localização']}]"
                    for _, row in man_opcoes.iterrows()
                ]
                escolha_man = st.selectbox("Selecione o item:", rotulos_man, key="man_selectbox")
                idx_man = rotulos_man.index(escolha_man)
                man_item_sel = man_opcoes.iloc[idx_man]
            else:
                st.warning("Nenhum item encontrado no estoque.")

        with col_qtd:
            man_quantidade = st.number_input("Quantidade", min_value=1, value=1, step=1, key="man_qtd")

        with col_un:
            man_un = st.selectbox(
                "UN",
                ["PC", "M", "KG", "UN", "CJ", "JG", "RL", "MT", "LT", "CX", "PAR", "KIT"],
                key="man_un"
            )

        if st.button("➕ Adicionar Item à Solicitação", use_container_width=True, key="btn_add_man"):
            if man_item_sel is None:
                st.error("Selecione um item do estoque antes de adicionar.")
            else:
                seq_num = len(st.session_state.req_manual_itens) + 1
                novo_item = {
                    "Seq": f"{seq_num:02d}",
                    "Localização": man_item_sel["Localização"],
                    "Código": man_item_sel["Código"],
                    "Descrição": man_item_sel["Descrição"],
                    "UN": man_un,
                    "Quantidade": float(man_quantidade),
                }
                st.session_state.req_manual_itens.append(novo_item)
                st.success(f"✅ Item **{man_item_sel['Descrição']}** adicionado.")

        st.markdown("---")

        # --- LISTA DE ITENS ATUAL ---
        st.subheader(f"📝 Lista de Itens ({len(st.session_state.req_manual_itens)} item(ns))")

        if st.session_state.req_manual_itens:
            df_man = pd.DataFrame(st.session_state.req_manual_itens)
            st.dataframe(df_man, use_container_width=True, hide_index=True)

            col_rem, col_limpar = st.columns(2)
            with col_rem:
                idx_remover = st.number_input(
                    "Nº Seq para remover:", min_value=1,
                    max_value=len(st.session_state.req_manual_itens), value=1,
                    step=1, key="man_idx_rem"
                )
                if st.button("❌ Remover item", key="btn_rem_man"):
                    st.session_state.req_manual_itens = [
                        it for it in st.session_state.req_manual_itens
                        if it["Seq"] != f"{idx_remover:02d}"
                    ]
                    for i, it in enumerate(st.session_state.req_manual_itens):
                        it["Seq"] = f"{i+1:02d}"
                    st.rerun()

            with col_limpar:
                if st.button("🗑️ Limpar toda a lista", key="btn_limpar_man"):
                    st.session_state.req_manual_itens = []
                    st.rerun()
        else:
            st.info("Nenhum item adicionado ainda.")

        st.markdown("---")

        # --- SALVAR / EXPORTAR ---
        st.subheader("💾 Salvar / Exportar")
        st.info(
            "ℹ️ O **Nº da Solicitação** é gerado automaticamente ao salvar, "
            "no formato: CódigoCliente + Data + Sequencial (ex: 10900904202601)."
        )
        col_sv1, col_sv2 = st.columns(2)

        def _montar_cabecalho_manual():
            return {
                "tipo_pedido": man_tipo_pedido,
                "cliente_cod": man_cliente_cod,
                "cliente_nome": man_cliente_nome,
                "comprador": man_comprador,
                "vendedor": man_vendedor_sel,
                "tipo_venda": man_tipo_venda,
                "departamento": man_departamento,
                "marcacoes": man_marcacoes,
                "observacao": man_observacao,
                "orcamento_pdf": "",
            }

        with col_sv1:
            if st.button("💾 Salvar Solicitação", use_container_width=True, key="btn_salvar_man"):
                if not st.session_state.req_manual_itens:
                    st.error("Adicione pelo menos um item antes de salvar.")
                elif not man_cliente_cod:
                    st.error("Selecione um cliente antes de salvar.")
                else:
                    cab_man = _montar_cabecalho_manual()
                    nome_salvo = salvar_requisicao(cab_man, st.session_state.req_manual_itens)
                    num_gerado = cab_man.get("num_solicitacao", "—")
                    st.success(f"✅ Solicitação salva! Nº: **{num_gerado}** | Arquivo: **{nome_salvo}**")

        with col_sv2:
            if st.session_state.req_manual_itens:
                cab_man_exp = _montar_cabecalho_manual()
                dados_man_excel = exportar_requisicao_excel(cab_man_exp, st.session_state.req_manual_itens)
                st.download_button(
                    label="📥 Exportar para Excel",
                    data=dados_man_excel,
                    file_name=f"SOL_{man_cliente_cod or 'nova'}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="btn_dl_man"
                )

    # =======================================================================
    # SUB-ABA: IMPORTAR VIA PDF
    # =======================================================================
    with sub_pdf:
        st.subheader("📄 Importar Solicitação via PDF")
        st.markdown(
            "Faça o upload de um arquivo PDF de requisição de materiais. "
            "O sistema irá **ler, interpretar e extrair** automaticamente os dados do cabeçalho e a lista de itens."
        )

        if not PDF_DISPONIVEL:
            st.error(
                "❌ A biblioteca **pdfplumber** não está instalada. "
                "Execute `pip install pdfplumber` no terminal e reinicie o aplicativo."
            )
        else:
            st.markdown("---")

            pdf_upload = st.file_uploader(
                "Selecione o arquivo PDF da requisição:",
                type=["pdf"],
                key="upload_req_pdf"
            )

            if pdf_upload is not None:
                with st.spinner("🔍 Lendo e interpretando o PDF..."):
                    try:
                        resultado = extrair_dados_requisicao(pdf_upload)
                        cabecalho = resultado["cabecalho"]
                        itens = resultado["itens"]
                        st.session_state.req_cabecalho = cabecalho
                        st.session_state.req_itens = itens
                        st.success(f"✅ PDF lido com sucesso! **{len(itens)}** item(ns) encontrado(s).")
                    except Exception as e:
                        st.error(f"❌ Erro ao processar o PDF: {e}")
                        cabecalho = {}
                        itens = []

                if cabecalho or itens:
                    st.markdown("---")
                    st.subheader("📌 Dados do Cabeçalho")
                    col_c1, col_c2 = st.columns(2)
                    with col_c1:
                        st.markdown(f"**Nº Orçamento PDF:** {cabecalho.get('orcamento_pdf', '—')}")
                        st.markdown(f"**Cliente:** {cabecalho.get('cliente_cod', '')} — {cabecalho.get('cliente_nome', '—')}")
                        st.markdown(f"**Vendedor:** {cabecalho.get('vendedor', '—')}")
                        st.markdown(f"**Comprador:** {cabecalho.get('comprador', '—')}")
                    with col_c2:
                        st.markdown(f"**Tipo de Venda:** {cabecalho.get('tipo_venda', '—')}")
                        st.markdown(f"**Departamento:** {cabecalho.get('departamento', '—')}")
                        st.markdown(f"**Marcações:** {cabecalho.get('marcacoes', '—')}")
                        st.markdown(f"**Observação:** {cabecalho.get('observacao', '—')}")

                    st.markdown("---")
                    st.subheader(f"📦 Itens Extraídos ({len(itens)} itens)")
                    if itens:
                        df_itens = pd.DataFrame(itens)
                        st.dataframe(df_itens, use_container_width=True, hide_index=True)
                    else:
                        st.warning("Nenhum item foi extraído do PDF. Verifique o formato do arquivo.")

                    st.markdown("---")
                    st.subheader("💾 Ações")
                    st.info(
                        "ℹ️ O **Nº da Solicitação** é gerado automaticamente ao salvar, "
                        "vinculando o Nº do Orçamento PDF ao sistema."
                    )
                    col_a1, col_a2 = st.columns(2)
                    with col_a1:
                        if st.button("💾 Salvar no Sistema", use_container_width=True, key="btn_salvar_req"):
                            if not itens:
                                st.error("Não há itens para salvar.")
                            else:
                                nome_salvo = salvar_requisicao(cabecalho, itens)
                                num_gerado = cabecalho.get("num_solicitacao", "—")
                                st.success(f"✅ Salvo! Nº Solicitação: **{num_gerado}** | Arquivo: **{nome_salvo}**")
                    with col_a2:
                        if itens:
                            dados_excel_req = exportar_requisicao_excel(cabecalho, itens)
                            st.download_button(
                                label="📥 Exportar para Excel",
                                data=dados_excel_req,
                                file_name=f"REQ_{cabecalho.get('orcamento_pdf', 'requisicao')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key="btn_download_req"
                            )

    # =======================================================================
    # SUB-ABA: SOLICITAÇÕES (HISTÓRICO AGRUPADO POR CLIENTE)
    # =======================================================================
    with sub_historico:
        st.subheader("📂 Solicitações")

        requisicoes_salvas = listar_requisicoes()

        if not requisicoes_salvas:
            st.info("Nenhuma solicitação salva ainda.")
        else:
            st.markdown(f"**{len(requisicoes_salvas)}** solicitação(ões) registrada(s).")
            st.markdown("---")

            # Agrupa por cliente
            clientes_dict: dict = {}
            for req in requisicoes_salvas:
                chave = f"{req['cliente_cod']} — {req['cliente_nome']}"
                if chave not in clientes_dict:
                    clientes_dict[chave] = []
                clientes_dict[chave].append(req)

            for cliente_label, reqs in sorted(clientes_dict.items()):
                st.markdown(f"### 👤 {cliente_label}  ·  {len(reqs)} solicitação(ões)")
                for req in reqs:
                    cab = req["cabecalho"]
                    titulo_exp = (
                        f"📄 Nº Sol.: {req['num_solicitacao']}  |  "
                        f"PDF: {req['orcamento_pdf']}  |  "
                        f"Tipo: {cab.get('tipo_pedido', cab.get('tipo_venda', '—'))}  |  "
                        f"Itens: {req['qtd_itens']}  |  "
                        f"Data: {req['data_importacao']}"
                    )
                    with st.expander(titulo_exp):
                        col_h1, col_h2 = st.columns(2)
                        with col_h1:
                            st.markdown(f"**Vendedor:** {cab.get('vendedor', '—')}")
                            st.markdown(f"**Comprador:** {cab.get('comprador', '—')}")
                            st.markdown(f"**Departamento:** {cab.get('departamento', '—')}")
                        with col_h2:
                            st.markdown(f"**Tipo de Venda:** {cab.get('tipo_venda', '—')}")
                            st.markdown(f"**Marcações:** {cab.get('marcacoes', '—')}")
                            st.markdown(f"**Observação:** {cab.get('observacao', '—')}")

                        if req["itens"]:
                            df_req_hist = pd.DataFrame(req["itens"])
                            st.dataframe(df_req_hist, use_container_width=True, hide_index=True)

                        dados_re_export = exportar_requisicao_excel(cab, req["itens"])
                        st.download_button(
                            label="📥 Exportar Excel",
                            data=dados_re_export,
                            file_name=f"SOL_{req['num_solicitacao']}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"re_export_{req['num_solicitacao']}"
                        )
                st.markdown("---")
