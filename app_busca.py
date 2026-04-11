# =============================================================================
#  Lúmen Bot — NEMA
#  A inteligência que acende a sua obra
#  + Processador de Chamados de Iluminação Pública
#  + Solicitação de Materiais (Manual e via PDF)
#  Setor de Obras e Instalações Elétricas — NEMA
#  Desenvolvido com Python + Streamlit
# =============================================================================

import os
import re
import json
import glob
import io
import subprocess
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

try:
    import bcrypt
    BCRYPT_OK = True
except ImportError:
    import hashlib
    BCRYPT_OK = False

try:
    import pdfplumber
    PDF_DISPONIVEL = True
except ImportError:
    PDF_DISPONIVEL = False

# ---------------------------------------------------------------------------
# Configuração global da página
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Lúmen Bot — NEMA",
    page_icon="💡",
    layout="wide",
)

# ---------------------------------------------------------------------------
# Constantes e caminhos
# ---------------------------------------------------------------------------
BASE_DIR          = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH        = os.path.join(BASE_DIR, "estoque.xlsx")
HISTORICO_DIR     = os.path.join(BASE_DIR, "historico_devolucoes")
REQUISICOES_DIR   = os.path.join(BASE_DIR, "historico_requisicoes")
RASCUNHO_PATH     = os.path.join(BASE_DIR, "rascunho_solicitacao.json")
USUARIOS_PATH     = os.path.join(BASE_DIR, "usuarios.json")

REF_CLIENTES      = os.path.join(BASE_DIR, "clientes.xlsx")
REF_VENDEDORES    = os.path.join(BASE_DIR, "vendedores.xlsx")
REF_COMPRADORES   = os.path.join(BASE_DIR, "compradores.xlsx")
REF_TIPOS_VENDA   = os.path.join(BASE_DIR, "tipos_venda.xlsx")
REF_DEPARTAMENTOS = os.path.join(BASE_DIR, "departamentos.xlsx")

os.makedirs(HISTORICO_DIR, exist_ok=True)
os.makedirs(REQUISICOES_DIR, exist_ok=True)


# ===========================================================================
# FUNÇÕES DE AUTENTICAÇÃO
# ===========================================================================

def _hash_senha(senha: str) -> str:
    if BCRYPT_OK:
        return bcrypt.hashpw(senha.encode(), bcrypt.gensalt()).decode()
    else:
        return hashlib.sha256(senha.encode()).hexdigest()


def _verificar_senha(senha: str, hash_salvo: str) -> bool:
    if BCRYPT_OK:
        try:
            return bcrypt.checkpw(senha.encode(), hash_salvo.encode())
        except Exception:
            return False
    else:
        return hashlib.sha256(senha.encode()).hexdigest() == hash_salvo


def carregar_usuarios() -> dict:
    """Carrega o arquivo de usuários. Cria admin padrão se não existir."""
    if not os.path.exists(USUARIOS_PATH):
        usuarios = {
            "admin": {
                "nome": "Administrador",
                "senha_hash": _hash_senha("Ameixaseca9988?"),
                "perfil": "admin",
                "ativo": True,
            }
        }
        salvar_usuarios(usuarios)
        return usuarios
    try:
        with open(USUARIOS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def salvar_usuarios(usuarios: dict):
    with open(USUARIOS_PATH, "w", encoding="utf-8") as f:
        json.dump(usuarios, f, ensure_ascii=False, indent=4)
    # Tenta fazer commit automático no Git para persistir no Streamlit Cloud
    _git_commit_file(USUARIOS_PATH, "Atualização de usuários")


def _git_commit_file(filepath: str, msg: str):
    """Faz git add + commit + push do arquivo para persistir no Streamlit Cloud."""
    try:
        cwd = BASE_DIR
        subprocess.run(["git", "add", filepath], cwd=cwd, capture_output=True, timeout=10)
        subprocess.run(["git", "commit", "-m", msg], cwd=cwd, capture_output=True, timeout=10)
        subprocess.run(["git", "push"], cwd=cwd, capture_output=True, timeout=30)
    except Exception:
        pass  # Falha silenciosa — funciona localmente sem Git


def autenticar(usuario: str, senha: str) -> bool:
    usuarios = carregar_usuarios()
    u = usuarios.get(usuario.strip().lower())
    if not u:
        return False
    if not u.get("ativo", True):
        return False
    return _verificar_senha(senha, u["senha_hash"])


def get_perfil(usuario: str) -> str:
    usuarios = carregar_usuarios()
    u = usuarios.get(usuario.strip().lower(), {})
    return u.get("perfil", "usuario")


# ===========================================================================
# TELA DE LOGIN
# ===========================================================================

def mostrar_login():
    # CSS completo da tela de login — design dark industrial
    st.markdown("""
    <style>
        /* Esconde sidebar e header padrão na tela de login */
        [data-testid="stSidebar"] { display: none !important; }
        header { display: none !important; }
        #MainMenu { display: none !important; }
        footer { display: none !important; }

        /* Fundo escuro com mesh gradient */
        .stApp {
            background-color: #121414 !important;
            background-image:
                radial-gradient(at 0% 0%, rgba(26, 35, 126, 0.15) 0px, transparent 50%),
                radial-gradient(at 100% 100%, rgba(255, 152, 0, 0.1) 0px, transparent 50%) !important;
        }
        .block-container {
            padding-top: 2rem !important;
            max-width: 420px !important;
            margin: 0 auto !important;
        }

        /* Barra gradiente no topo */
        .top-gradient-bar {
            position: fixed; top: 0; left: 0; width: 100%; height: 4px; z-index: 9999;
            background: linear-gradient(to right, #1a237e, #ff9800, #1a237e);
            opacity: 0.3;
        }

        /* Labels dos inputs */
        label {
            color: #908f9d !important;
            font-size: 0.65rem !important;
            font-weight: 700 !important;
            letter-spacing: 0.15em !important;
            text-transform: uppercase !important;
        }

        /* Inputs escuros */
        input[type="text"], input[type="password"] {
            background-color: #282a2b !important;
            border: none !important;
            border-radius: 8px !important;
            color: #e2e2e2 !important;
            font-size: 0.9rem !important;
            padding: 0.9rem 1rem !important;
        }
        input[type="text"]:focus, input[type="password"]:focus {
            box-shadow: 0 0 0 1px rgba(255,152,0,0.3) !important;
        }
        input::placeholder {
            color: rgba(144,143,157,0.5) !important;
        }

        /* Botão de login — gradiente laranja */
        div[data-testid="stForm"] button[kind="primaryFormSubmit"],
        div[data-testid="stForm"] button {
            background: linear-gradient(to right, #ff9800, #ffb870) !important;
            color: #4a2800 !important;
            font-weight: 700 !important;
            font-size: 0.8rem !important;
            letter-spacing: 0.15em !important;
            text-transform: uppercase !important;
            border-radius: 12px !important;
            border: none !important;
            width: 100% !important;
            padding: 0.9rem !important;
            box-shadow: 0 4px 15px rgba(255,152,0,0.25) !important;
            transition: all 0.2s !important;
        }
        div[data-testid="stForm"] button:hover {
            box-shadow: 0 6px 20px rgba(255,152,0,0.35) !important;
        }
        div[data-testid="stForm"] button:active {
            transform: scale(0.97) !important;
        }

        /* Mensagem de erro */
        .login-erro {
            color: #ffb4ab;
            background-color: rgba(147,0,10,0.4);
            border-radius: 8px;
            padding: 0.6rem 1rem;
            text-align: center;
            margin-top: 0.5rem;
            font-weight: 600;
            font-size: 0.85rem;
        }
    </style>
    <div class="top-gradient-bar"></div>
    """, unsafe_allow_html=True)

    # ---- Ícone da lâmpada + Branding ----
    st.markdown("""
    <div style="text-align:center; margin-bottom: 2.5rem;">
        <div style="display:inline-flex; align-items:center; justify-content:center;
                    width:80px; height:80px; border-radius:12px;
                    background-color:#282a2b;
                    box-shadow: 0 0 32px rgba(226,226,226,0.04);
                    margin-bottom: 1rem;">
            <span style="font-size:2.8rem;">💡</span>
        </div>
        <h1 style="color:#e2e2e2; font-size:1.9rem; font-weight:800;
                   letter-spacing:-0.03em; text-transform:uppercase;
                   margin:0; font-family:'Inter',sans-serif;">
            Lumen Control
        </h1>
        <p style="color:#908f9d; font-size:0.65rem; letter-spacing:0.2em;
                  text-transform:uppercase; margin:0.3rem 0 0 0;
                  font-weight:500;">
            NEMA &bull; Acesso Seguro
        </p>
    </div>
    """, unsafe_allow_html=True)

    # ---- Formulário de login ----
    with st.form("form_login", clear_on_submit=False):
        usuario_input = st.text_input("Node Identity", placeholder="Usuário", key="login_usuario")
        senha_input   = st.text_input("Access Key", type="password", placeholder="••••••••", key="login_senha")
        st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
        entrar = st.form_submit_button("Iniciar Sessão", use_container_width=True)

    if entrar:
        if autenticar(usuario_input, senha_input):
            st.session_state.logado = True
            st.session_state.usuario_logado = usuario_input.strip().lower()
            st.session_state.perfil_logado = get_perfil(usuario_input.strip().lower())
            st.session_state.pagina_atual = "🏠 Início"
            st.rerun()
        else:
            st.markdown(
                "<div class='login-erro'>❌ Usuário ou senha incorretos.</div>",
                unsafe_allow_html=True
            )

    # ---- Rodapé ----
    st.markdown("""
    <div style="text-align:center; margin-top:3rem; opacity:0.4;">
        <div style="width:48px; height:4px; background:#454652; border-radius:9999px;
                    margin:0 auto 1rem auto;"></div>
        <p style="font-size:0.55rem; text-transform:uppercase; letter-spacing:0.2em;
                  font-weight:700; color:#908f9d;">
            Lumen Systems &bull; NEMA Tecnologia &bull; Rio do Sul, SC
        </p>
    </div>
    """, unsafe_allow_html=True)


# ===========================================================================
# FUNÇÕES AUXILIARES — APP PRINCIPAL
# ===========================================================================

@st.cache_data
def carregar_estoque() -> pd.DataFrame:
    df = pd.read_excel(EXCEL_PATH, dtype=str)
    for col in ["Código", "Descrição", "Localização"]:
        if col not in df.columns:
            st.error(f"Coluna '{col}' não encontrada no arquivo estoque.xlsx.")
            st.stop()
    df["Código"] = df["Código"].str.strip()
    df["Descrição"] = df["Descrição"].str.strip()
    df["Localização"] = df["Localização"].str.strip()
    df.dropna(how="all", inplace=True)
    df.sort_values("Descrição", inplace=True, ignore_index=True)
    return df


@st.cache_data
def carregar_referencia(caminho: str, colunas: list) -> list:
    if not os.path.exists(caminho):
        return []
    try:
        df = pd.read_excel(caminho, dtype=str).fillna("")
        resultado = []
        for _, row in df.iterrows():
            partes = [str(row[c]).strip() for c in colunas if c in df.columns and str(row[c]).strip()]
            if partes:
                resultado.append(" — ".join(partes) if len(partes) > 1 else partes[0])
        return resultado
    except Exception:
        return []


@st.cache_data
def carregar_clientes() -> list:
    return carregar_referencia(REF_CLIENTES, ["Código", "Nome"])


@st.cache_data
def carregar_vendedores() -> list:
    return carregar_referencia(REF_VENDEDORES, ["Código", "Nome"])


@st.cache_data
def carregar_compradores() -> list:
    return carregar_referencia(REF_COMPRADORES, ["Nome"])


@st.cache_data
def carregar_tipos_venda() -> list:
    return carregar_referencia(REF_TIPOS_VENDA, ["Tipo"])


@st.cache_data
def carregar_departamentos() -> list:
    return carregar_referencia(REF_DEPARTAMENTOS, ["Código", "Nome"])


def gerar_protocolo() -> str:
    hoje = datetime.now().strftime("%d%m%Y")
    padrao = os.path.join(HISTORICO_DIR, f"DEV_{hoje}*.json")
    arquivos_hoje = glob.glob(padrao)
    sequencial = len(arquivos_hoje) + 1
    return f"{hoje}.{sequencial:03d}"


def gerar_numero_solicitacao(cliente_cod: str) -> str:
    hoje = datetime.now().strftime("%d%m%Y")
    cod = re.sub(r'\D', '', str(cliente_cod))
    prefixo = f"{cod}{hoje}"
    padrao = os.path.join(REQUISICOES_DIR, f"SOL_{prefixo}*.json")
    existentes = glob.glob(padrao)
    sequencial = len(existentes) + 1
    return f"{prefixo}{sequencial:02d}"


def df_lista_vazio() -> pd.DataFrame:
    return pd.DataFrame(columns=["Código", "Descrição", "Localização", "Quantidade"])


def salvar_rascunho(cabecalho: dict, itens: list):
    dados = {"cabecalho": cabecalho, "itens": itens}
    with open(RASCUNHO_PATH, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=4)


def carregar_rascunho() -> dict:
    if os.path.exists(RASCUNHO_PATH):
        try:
            with open(RASCUNHO_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def limpar_rascunho():
    if os.path.exists(RASCUNHO_PATH):
        os.remove(RASCUNHO_PATH)


def inicializar_session_state():
    if "logado" not in st.session_state:
        st.session_state.logado = False
    if "usuario_logado" not in st.session_state:
        st.session_state.usuario_logado = ""
    if "perfil_logado" not in st.session_state:
        st.session_state.perfil_logado = "usuario"
    if "lista_devolucao" not in st.session_state:
        st.session_state.lista_devolucao = df_lista_vazio()
    if "empresa" not in st.session_state:
        st.session_state.empresa = ""
    if "protocolo" not in st.session_state:
        st.session_state.protocolo = gerar_protocolo()
    if "req_cabecalho" not in st.session_state:
        st.session_state.req_cabecalho = {}
    if "req_itens" not in st.session_state:
        st.session_state.req_itens = []
    if "req_manual_itens" not in st.session_state:
        rascunho = carregar_rascunho()
        st.session_state.req_manual_itens = rascunho.get("itens", [])
    if "req_manual_cab" not in st.session_state:
        rascunho = carregar_rascunho()
        st.session_state.req_manual_cab = rascunho.get("cabecalho", {})
    if "pagina_atual" not in st.session_state:
        st.session_state.pagina_atual = "🏠 Início"


# ---------------------------------------------------------------------------
# Funções auxiliares — DEVOLUÇÃO
# ---------------------------------------------------------------------------

def salvar_devolucao():
    protocolo = st.session_state.protocolo
    dados = {
        "protocolo": protocolo,
        "empresa": st.session_state.empresa,
        "data_hora": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "itens": st.session_state.lista_devolucao.to_dict(orient="records"),
    }
    caminho = os.path.join(HISTORICO_DIR, f"DEV_{protocolo}.json")
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=4)
    st.success(f"Devolução salva com sucesso! Protocolo: **{protocolo}**")


def nova_devolucao():
    st.session_state.lista_devolucao = df_lista_vazio()
    st.session_state.empresa = ""
    st.session_state.protocolo = gerar_protocolo()


def exportar_excel_devolucao() -> bytes:
    df_export = st.session_state.lista_devolucao.copy()
    df_export.insert(0, "Protocolo", st.session_state.protocolo)
    df_export.insert(1, "Empresa", st.session_state.empresa)
    df_export["Data/Hora"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Devolução")
    return buffer.getvalue()


def listar_historico() -> list:
    arquivos = sorted(glob.glob(os.path.join(HISTORICO_DIR, "DEV_*.json")), reverse=True)
    registros = []
    for arq in arquivos:
        with open(arq, "r", encoding="utf-8") as f:
            dados = json.load(f)
        registros.append({
            "arquivo": arq,
            "protocolo": dados.get("protocolo", "—"),
            "empresa": dados.get("empresa", "—"),
            "data_hora": dados.get("data_hora", "—"),
            "qtd_itens": len(dados.get("itens", [])),
            "itens": dados.get("itens", []),
        })
    return registros


# ---------------------------------------------------------------------------
# Funções auxiliares — PROCESSADOR DE CHAMADOS IP
# ---------------------------------------------------------------------------

ROUTES = {
    "ROTA 1": ["CENTRO", "JARDIM AMERICA"],
    "ROTA 2": ["ALBERTINA", "LARANJEIRAS", "BOA VISTA", "EUGENIO SCHNEIDER"],
    "ROTA 3": ["FUNDO CANOAS", "CANOAS", "PROGRESSO", "PAMPLONA", "CANTA GALO"],
    "ROTA 4": ["BELA VISTA", "VILA NOVA", "BAIRRO NOVO", "JARDIM ESPERANÇA"],
    "ROTA 5": ["INDUSTRIAL", "JARDIM INDUSTRIAL", "ZONA INDUSTRIAL"],
}


def processar_chamados(
    uploaded_file, fonte_escolhida,
    cor_fundo_rota, cor_fonte_rota, tamanho_rota,
    cor_fundo_bairro, cor_fonte_bairro, tamanho_bairro,
    cor_fundo_prob, cor_fonte_prob, tamanho_prob,
) -> bytes:
    xls = pd.read_excel(uploaded_file, sheet_name=None, header=None, dtype=str)
    abas_disponiveis = {nome.strip().upper(): nome for nome in xls.keys()}
    data_by_route = {r: {} for r in ROUTES}

    for route, neighborhoods in ROUTES.items():
        for neighborhood in neighborhoods:
            nome_aba_upper = neighborhood.upper()
            if nome_aba_upper in abas_disponiveis:
                nome_real_aba = abas_disponiveis[nome_aba_upper]
                df = xls[nome_real_aba]
                if len(df.columns) >= 4:
                    df_filtered = df[
                        (df[3].isin(['NÃO REALIZADO', 'NÃO EXECUTADO'])) &
                        (df[1].notna()) & (df[1].str.strip() != "")
                    ]
                    problems = df_filtered[1].tolist()
                    if problems:
                        data_by_route[route][neighborhood] = problems

    wb = Workbook()
    ws = wb.active
    ws.title = "Chamados Pendentes"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    hex_bg_rota   = cor_fundo_rota.replace('#', '')
    hex_fg_rota   = cor_fonte_rota.replace('#', '')
    hex_bg_bairro = cor_fundo_bairro.replace('#', '')
    hex_fg_bairro = cor_fonte_bairro.replace('#', '')
    hex_bg_prob   = cor_fundo_prob.replace('#', '')
    hex_fg_prob   = cor_fonte_prob.replace('#', '')

    font_rota   = Font(name=fonte_escolhida, size=tamanho_rota,   color=hex_fg_rota,   bold=True)
    fill_rota   = PatternFill(start_color=hex_bg_rota,   end_color=hex_bg_rota,   fill_type="solid")
    font_bairro = Font(name=fonte_escolhida, size=tamanho_bairro, color=hex_fg_bairro, bold=True)
    fill_bairro = PatternFill(start_color=hex_bg_bairro, end_color=hex_bg_bairro, fill_type="solid")
    font_prob   = Font(name=fonte_escolhida, size=tamanho_prob,   color=hex_fg_prob)
    fill_prob   = PatternFill(start_color=hex_bg_prob,   end_color=hex_bg_prob,   fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

 
